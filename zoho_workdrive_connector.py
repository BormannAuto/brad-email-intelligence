"""
zoho_workdrive_connector.py
Bormann Marketing — Email Intelligence System v3  (NEW)
Handles Zoho WorkDrive API: auth, folder listing, file download and text extraction.
Scope: WorkDrive.files.READ

If env vars are not set, returns None gracefully — WorkDrive features disabled.
"""

import os
import io
import json
import logging
import requests
from typing import Optional

logger = logging.getLogger(__name__)

ZOHO_ACCOUNTS_URL    = "https://accounts.zoho.com/oauth/v2/token"
ZOHO_WORKDRIVE_BASE  = "https://workdrive.zoho.com/api/v1"


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

def authenticate_workdrive() -> Optional[dict]:
    """
    Load WorkDrive OAuth credentials from environment variables and refresh token.
    If env vars are not set, returns None — WorkDrive features are simply disabled.
    Returns session dict: {access_token} or None.
    """
    client_id     = os.environ.get("ZOHO_WORKDRIVE_CLIENT_ID")
    client_secret = os.environ.get("ZOHO_WORKDRIVE_CLIENT_SECRET")
    refresh_token = os.environ.get("ZOHO_WORKDRIVE_REFRESH_TOKEN")

    # Graceful disable if not configured
    if not all([client_id, client_secret, refresh_token]):
        logger.info(
            "WorkDrive env vars not set — WorkDrive features disabled for this run."
        )
        return None

    try:
        resp = requests.post(ZOHO_ACCOUNTS_URL, data={
            "refresh_token": refresh_token,
            "client_id":     client_id,
            "client_secret": client_secret,
            "grant_type":    "refresh_token",
        }, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        if "access_token" not in data:
            logger.warning(f"WorkDrive token refresh returned no access_token: {data}")
            return None

        logger.info("Zoho WorkDrive authenticated successfully.")
        return {"access_token": data["access_token"]}

    except Exception as e:
        logger.warning(f"WorkDrive authentication failed: {e} — features disabled.")
        return None


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _wd_get(session: dict, path: str, stream: bool = False, params: Optional[dict] = None):
    """Authenticated GET against Zoho WorkDrive API."""
    url     = f"{ZOHO_WORKDRIVE_BASE}/{path}"
    headers = {
        "Authorization": f"Zoho-oauthtoken {session['access_token']}",
        "Accept":        "application/json",
    }
    resp = requests.get(url, headers=headers, params=params or {}, stream=stream, timeout=60)
    resp.raise_for_status()
    return resp


# ---------------------------------------------------------------------------
# Config loader
# ---------------------------------------------------------------------------

def _get_brand_folder_map() -> dict:
    """Load workdrive_brand_folders map from email_config.json."""
    try:
        with open("email_config.json") as f:
            config = json.load(f)
        return config.get("workdrive_brand_folders", {})
    except Exception as e:
        logger.warning(f"Could not load workdrive_brand_folders from email_config.json: {e}")
        return {}


# ---------------------------------------------------------------------------
# Public functions
# ---------------------------------------------------------------------------

def list_brand_folder(session: Optional[dict], brand_name: str) -> list[dict]:
    """
    Look up folder ID from email_config.json workdrive_brand_folders map.
    Returns list of {file_id, file_name, file_type} in that folder.
    Returns empty list if session is None or brand not found.
    """
    if session is None:
        return []

    folder_map = _get_brand_folder_map()
    folder_id  = folder_map.get(brand_name) or folder_map.get(brand_name.lower())
    if not folder_id:
        logger.debug(f"No WorkDrive folder configured for brand '{brand_name}'")
        return []

    try:
        resp  = _wd_get(session, f"files/{folder_id}/files")
        data  = resp.json()
        files = data.get("data", [])
        result = []
        for f in files:
            attrs      = f.get("attributes", {})
            file_type  = attrs.get("type", "").lower()
            # Only process supported types
            if file_type in ("pdf", "xlsx", "xls", "csv", "text"):
                result.append({
                    "file_id":   f.get("id", ""),
                    "file_name": attrs.get("name", ""),
                    "file_type": file_type,
                })
        logger.info(f"WorkDrive: found {len(result)} processable files for brand '{brand_name}'")
        return result

    except Exception as e:
        logger.warning(f"list_brand_folder failed for brand '{brand_name}': {e}")
        return []


def fetch_file_as_text(
    session: Optional[dict],
    file_id: str,
    file_type: str,
) -> Optional[str]:
    """
    Download file and extract as plain text.
    PDF: uses pdfplumber. xlsx/xls: uses openpyxl. csv: decoded directly.
    Returns plain text string, or None on failure.
    """
    if session is None:
        return None

    try:
        # Download the file
        resp = _wd_get(session, f"download/{file_id}", stream=True)
        raw_bytes = resp.content

        if file_type == "pdf":
            return _extract_pdf_text(raw_bytes)
        elif file_type in ("xlsx", "xls"):
            return _extract_excel_text(raw_bytes)
        elif file_type == "csv":
            return raw_bytes.decode("utf-8", errors="replace")
        else:
            # Attempt plain text decode
            return raw_bytes.decode("utf-8", errors="replace")

    except Exception as e:
        logger.error(f"fetch_file_as_text failed for file {file_id} ({file_type}): {e}")
        return None


def _extract_pdf_text(raw_bytes: bytes) -> Optional[str]:
    """Extract text from PDF bytes using pdfplumber."""
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        return "\n\n".join(text_parts)
    except ImportError:
        logger.error("pdfplumber not installed — cannot extract PDF text.")
        return None
    except Exception as e:
        logger.error(f"PDF text extraction failed: {e}")
        return None


def _extract_excel_text(raw_bytes: bytes) -> Optional[str]:
    """Extract Excel content as plain text using openpyxl."""
    try:
        import openpyxl
        wb     = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        parts  = []
        for sheet in wb.sheetnames:
            ws   = wb[sheet]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                if row_text.strip():
                    rows.append(row_text)
            if rows:
                parts.append(f"[Sheet: {sheet}]\n" + "\n".join(rows))
        return "\n\n".join(parts)
    except ImportError:
        logger.error("openpyxl not installed — cannot extract Excel text.")
        return None
    except Exception as e:
        logger.error(f"Excel text extraction failed: {e}")
        return None
